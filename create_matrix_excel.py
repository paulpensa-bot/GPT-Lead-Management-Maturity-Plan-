#!/usr/bin/env python3
"""
Create Lead Management Maturity Matrix Excel file
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Maturity Matrix"

# Define colors
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
workstream_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
stage1_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
stage2_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
stage3_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
stage4_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
stage5_fill = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")

# Define borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 30
for col in range(2, 7):
    ws.column_dimensions[get_column_letter(col)].width = 45

# Title
ws['A1'] = "Lead Management Maturity Matrix"
ws.merge_cells('A1:F1')
ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].fill = header_fill
ws.row_dimensions[1].height = 30

# Headers
ws['A2'] = "Workstream"
ws['B2'] = "Stage 1: Documented"
ws['C2'] = "Stage 2: Operationally Consistent"
ws['D2'] = "Stage 3: Enterprise Integrated"
ws['E2'] = "Stage 4: Quantitatively Managed"
ws['F2'] = "Stage 5: Self-Optimizing"

for col in range(1, 7):
    cell = ws.cell(row=2, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.fill = workstream_fill
    cell.border = thin_border

ws.row_dimensions[2].height = 40

# Data for each workstream
workstreams = [
    {
        "name": "1. Lead Strategy & Execution",
        "stages": [
            "• Basic workflows documented\n• Draft SLAs exist\n• Manual routing\n• Limited cross-channel coordination",

            "• Operationalize SLAs (finalize values, configure CRM tracking, communicate to all channels)\n• Launch enterprise lead-routing policies (define rules by segment, create re-routing rules, introduce fallback routing)\n• Establish cross-channel governance (weekly SLA compliance reviews, executive escalation path)\n• Introduce lead tiering/prioritization (define urgency levels, create differentiated SLAs, update CRM views)",

            "• Standardize workflows across ALL channels (create single enterprise workflow, eliminate process variance, publish v3.0)\n• Deploy Enterprise Routing Engine v2 (multi-factor routing with client value/lead score/licensing/capacity)\n• Stand up enterprise governance with KPIs & consequences (monthly reviews of conversion/routing/data quality/SLA by channel)",

            "• Implement Dynamic Orchestration Engine (ML-assisted routing using propensity models, advisor capacity, engagement data)\n• Automate lead lifecycle transitions (auto-update statuses, remove manual steps, create multi-step workflows)\n• Introduce enterprise lead forecasting (predict volume by channel/trigger/segment, adjust staffing)\n• Establish enterprise OKRs (conversion, lead effectiveness, SLA compliance, revenue attribution)",

            "• Deploy Autonomous Routing Engine v4 (reinforcement learning, self-adjusting weights, auto-experimentation)\n• Implement adaptive strategies (auto-adjust cadence/channel/prioritization, predict optimal sequences)\n• Introduce guardrail-based governance (automated monitoring, auto-escalations at control limits)\n• Connect strategy to OKRs (project lead-driven revenue, auto-adjust strategies, surface risk alerts)"
        ]
    },
    {
        "name": "2. Lead Creation",
        "stages": [
            "• Basic trigger library\n• Manual suppression\n• No lead scoring\n• Inconsistent hand-offs",

            "• Expand Trigger Library (add rollover propensity, investment drift, distribution initiation, contribution changes)\n• Automate suppression logic (turn rules into automation, add frequency capping, prevent overlap)\n• Introduce lead-quality scoring v1 (define inputs, create rules-based logic, add to CRM)\n• Standardize Marketing→CRM hand-off (enforce required fields, validate integrity, automate ingestion)",

            "• Expand Trigger Library v2→v3 (add web events, plan interactions, calculators, life events; normalize metadata)\n• Introduce Lead Scoring v2 (weighted model with account size, trigger type, engagement, past behavior)\n• Industrialize suppression (connect datasets across wealth/retirement/marketing/call center, build API/data layer)\n• Introduce cross-channel deduping (fuzzy matching, merge before routing, canonical identifier)",

            "• Build Predictive Trigger Models v4 (ML models for high-ROI triggers using behavioral patterns, plan risk, conversion history)\n• Real-time data pipeline (move from batch to event streaming, near-instant triggers with suppression)\n• Lead Scoring v3 Predictive (ML model per business line, dynamic score updates based on actions)\n• Full-funnel suppression orchestration (span all channels, add explanations, AI-based over-touch prevention)",

            "• Deploy self-optimizing triggers (ML updates own weights, system creates/retires triggers, unsupervised learning)\n• Event-driven creation at scale (instant triggers from risk changes, market shocks, anomalies; composite triggers)\n• Implement adaptive suppression (AI predicts fatigue, auto-adjusts frequency, personalized by preference)\n• Connect to LTV models (prioritize by long-term value, drop low-LTV triggers automatically)"
        ]
    },
    {
        "name": "3. Salesforce CRM Enhancements",
        "stages": [
            "• Structured fields\n• Basic reporting\n• Manual data entry\n• No validation rules",

            "• Add enforcement layers (make SLA fields required, ensure complete lead data, prevent incorrect dispositions)\n• Enhance routing automation (build deterministic routing by client type/trigger/wealth threshold/capacity, add re-routing for missed SLAs)\n• Improve CRM UX Stage 2 (create simplified workspace with SLA timers, prioritized queue, trigger context panel)\n• Establish CRM adoption monitoring (weekly hygiene reports, enforce disposition standards, identify non-compliant users)",

            "• Deploy Lead Workspace v2 (add lead score, trigger context, SLA timer, recommended actions panels)\n• Implement advanced workflow automation (auto-create tasks, recycle unworked leads, trigger SLA breach notifications)\n• Integrate with marketing/retirement/advisor systems (build APIs for digital events, trigger data, book-of-business context)\n• Establish usage scorecards (track data hygiene, disposition accuracy, SLA compliance, activity logs by user)",

            "• Deploy AI Next Best Action (provide recommended outreach based on profile/history/similar outcomes)\n• AI-generated call summaries (auto-summarize calls/emails/chats, tag sentiment, identify opportunities)\n• Real-time SLA monitoring & enforcement (alerts for risks, auto-route if breach imminent, auto-escalations)\n• Unified Retirement+Wealth 360° view (merge plan/household data, integrate opportunities, 1-click segmentation)",

            "• Deploy Agentic CRM workflows (AI drafts outreach, schedules follow-ups, prepares briefs, updates dispositions)\n• Full CRM→Data Cloud fusion (unified customer graph, sub-second queries, enriched with historical/predictive context)\n• CRM-driven adaptive workflows (modify paths based on behavior, route based on strengths, auto-tune UI)\n• Introduce Advisor Digital Twin (AI model of skills/habits/performance, predict ideal leads/coaching needs)"
        ]
    },
    {
        "name": "4. Measurement & Reporting",
        "stages": [
            "• Basic dashboards\n• Limited KPIs\n• No SLA tracking\n• Siloed by channel",

            "• Expand dashboards for SLA measurement (track acceptance/first-attempt/cadence, display breach flags, create per-channel funnel)\n• Build cross-channel reporting (combine Retirement+Wealth, create single enterprise funnel, normalize dispositions)\n• Implement weekly data quality monitoring (detect missing fields/invalid dispositions/duplicates, produce DQ scorecard)\n• Introduce early attribution models (assign leads to channels, track revenue influence, prepare for multi-touch)",

            "• Build unified enterprise funnel dashboard (combine funnels, add drop-off/leakage/root-cause analysis, add segmentation views)\n• Implement closed-loop performance reporting (capture trigger→routed→worked→converted, build lead→revenue attribution)\n• Mature Data Dictionary v2→v3 (standardize across businesses, add metadata fields, create data contracts)\n• Introduce KPI targets & benchmarking (establish targets for acceptance/attempts/disposition/conversion; compare teams)",

            "• Implement real-time dashboards (streaming data updates for SLA, conversion windows, attribution, advisor productivity)\n• Build predictive conversion models (predict which leads convert and why, forecast by trigger/segment/persona)\n• Launch automated attribution engine (multi-touch across email/CRM/calls/digital, connect to revenue, produce ROI dashboards)\n• Introduce operational control limits (identify statistical norms, set acceptable ranges, trigger alerts on drift)",

            "• Deploy prescriptive analytics (recommend actions based on patterns/leakage/capacity, propose routing/trigger/playbook changes)\n• Autonomous KPI management (system takes action on drift: adjust routing, increase coaching, modify rules)\n• Automated experimentation engine (AI runs continuous A/B tests on routing/triggers/cadences/workflows, chooses winners)\n• Financial contribution modeling (full lead→revenue→LTV→P&L linkage, real-time business cases, surface financial risks)"
        ]
    },
    {
        "name": "5. Practice Management",
        "stages": [
            "• Playbooks documented\n• Ad-hoc training\n• No capacity modeling\n• Inconsistent coaching",

            "• Roll out mandatory training (train all advisors/wholesalers/reps, test understanding, certify completion)\n• Implement coaching cadence (managers review weekly performance, coach on SLA/quality, introduce scorecards)\n• Introduce capacity modeling basic (document advisor capacity, adjust routing by availability, publish monthly projections)\n• Launch QA program (review recordings/emails/notes, score against playbooks, feed insights to managers/analytics)",

            "• Implement structured coaching across channels (weekly coaching per advisor/rep, create scripts aligned to metrics, tie to performance mgmt)\n• Deploy capacity modeling v2 predictive (account for seasonal volume, advisor pipeline, workload variability)\n• Standardize performance scorecards by persona (identical KPIs for advisor/rep/wholesaler, create tiered groups, tie to compensation)\n• Formalize QA with enterprise rubric (standard rubric across businesses, review 5-10 interactions/month, feed to coaching loop)",

            "• Introduce performance-based routing (use QA/SLA/conversion to influence routing, reward top performers, reduce load for coaching needs)\n• Implement AI-enabled coaching (use call summaries for opportunities, predict drift/overload, recommend training modules)\n• Institutionalize advanced capacity planning (predict staffing weeks ahead, plan for surges, feed into routing engine)\n• Redesign compensation (add incentives for SLA/hygiene/conversion quality, introduce penalties for non-compliance)",

            "• AI-driven workforce optimization (real-time allocation based on demand, workload simulation, dynamic staffing suggestions)\n• Personalized AI coaching (customized micro-learning, identify root causes, on-demand vs scheduled)\n• Advisor mastery tiers auto-updated (system assigns tier, higher tiers get complex leads, prescriptive steps to advance)\n• Transform to dynamic rewards engine (compensation auto-adjusts to tiers, tied to OKRs, bonuses for AI-identified behaviors)"
        ]
    }
]

# Fill in the workstreams and stages
current_row = 3
stage_fills = [stage1_fill, stage2_fill, stage3_fill, stage4_fill, stage5_fill]

for workstream in workstreams:
    # Workstream name
    ws.cell(row=current_row, column=1, value=workstream["name"])
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=current_row, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws.cell(row=current_row, column=1).border = thin_border

    # Stage data
    for col_idx, stage_text in enumerate(workstream["stages"], start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=stage_text)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = thin_border
        cell.fill = stage_fills[col_idx - 2]
        cell.font = Font(size=10)

    # Set row height
    ws.row_dimensions[current_row].height = 150
    current_row += 1

# Add summary sheet
ws_summary = wb.create_sheet("Stage Descriptions")
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 80

# Summary title
ws_summary['A1'] = "Maturity Stage Descriptions"
ws_summary.merge_cells('A1:B1')
ws_summary['A1'].font = Font(size=14, bold=True, color="FFFFFF")
ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_summary['A1'].fill = header_fill
ws_summary.row_dimensions[1].height = 25

summary_data = [
    ["Stage 1: Documented", "Foundation level with basic processes documented but inconsistently followed. Manual operations dominate. SLAs exist but aren't enforced, routing is manual, and cross-channel coordination is limited."],
    ["Stage 2: Operationally Consistent", "Processes are enforced with SLAs, basic automation begins, cross-channel alignment established, and governance introduced. Systems enforce data quality and routing becomes semi-automated with fallback logic."],
    ["Stage 3: Enterprise Integrated", "Standardized workflows across all channels, unified systems, comprehensive measurement, and enterprise-wide governance. Single enterprise workflow replaces siloed processes, and routing becomes dynamic based on multiple factors."],
    ["Stage 4: Quantitatively Managed", "Automated workflows with AI assistance, predictive analytics, real-time measurement, and data-driven decision making. Machine learning drives routing, scoring, and recommendations. Real-time dashboards with predictive models."],
    ["Stage 5: Self-Optimizing", "Autonomous systems that learn and adapt, prescriptive analytics, continuous improvement embedded, and intelligent orchestration. Reinforcement learning optimizes routing automatically, AI performs tasks not just suggests, and the system self-heals performance issues."]
]

current_row = 3
for stage_name, stage_desc in summary_data:
    ws_summary.cell(row=current_row, column=1, value=stage_name)
    ws_summary.cell(row=current_row, column=1).font = Font(bold=True, size=11)
    ws_summary.cell(row=current_row, column=1).alignment = Alignment(vertical='top', wrap_text=True)
    ws_summary.cell(row=current_row, column=1).fill = workstream_fill
    ws_summary.cell(row=current_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws_summary.cell(row=current_row, column=1).border = thin_border

    ws_summary.cell(row=current_row, column=2, value=stage_desc)
    ws_summary.cell(row=current_row, column=2).alignment = Alignment(vertical='top', wrap_text=True)
    ws_summary.cell(row=current_row, column=2).border = thin_border

    ws_summary.row_dimensions[current_row].height = 60
    current_row += 1

# Save the workbook
wb.save('/home/user/GPT-Lead-Management-Maturity-Plan-/Lead-Management-Maturity-Matrix.xlsx')
print("Excel file created successfully!")
